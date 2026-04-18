"""
Sky Engineering Team Registry — Data Population Command
Author: Maurya Patel (Lead Developer, Student 4)

Populates the database with REAL data from the official Sky Engineering
Team Registry Excel spreadsheet.
"""

import os
import openpyxl
from datetime import time, datetime, timedelta
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from core.models import (
    Department, Team, TeamMember, Dependency,
    ContactChannel, StandupInfo, RepositoryLink,
    WikiLink, BoardLink, AuditLog
)

class Command(BaseCommand):
    help = 'Populates the database with real Sky Engineering team data from Excel.'

    def handle(self, *args, **kwargs):
        self.stdout.write(self.style.WARNING('\n Starting Sky Registry data population from Excel...\n'))

        excel_path = os.path.join(
            'only_for_to_read_context_no_github_push',
            'Agile Project Module UofW - Team Registry.xlsx'
        )

        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f' Excel file not found at: {excel_path}'))
            return

        with transaction.atomic():
            # Clear existing data to avoid duplicates and ensure a clean registry
            BoardLink.objects.all().delete()
            WikiLink.objects.all().delete()
            RepositoryLink.objects.all().delete()
            StandupInfo.objects.all().delete()
            ContactChannel.objects.all().delete()
            Dependency.objects.all().delete()
            TeamMember.objects.all().delete()
            AuditLog.objects.all().delete()
            Team.objects.all().delete()
            Department.objects.all().delete()
            self.stdout.write('  Cleared existing registry data')

            # Load Workbook
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active

            # Departments mapping (standardised for Sky)
            # We now derive leaders directly from the Excel 'Department Head' column
            depts = {}
            
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            
            # First pass: Create unique departments from the sheet
            for row in rows:
                if not row or not row[0]: continue
                dept_name = row[0]
                dept_head = row[2] or "TBC"
                
                if dept_name not in depts:
                    dept = Department.objects.get_or_create(
                        department_name=dept_name,
                        defaults={
                            'department_lead_name': dept_head,
                            'specialization': "Engineering",
                            'description': f"Official Sky Engineering Department: {dept_name}"
                        }
                    )[0]
                    depts[dept_name] = dept

            # Process Teams
            teams_created = {}
            
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[3]: # Skip if no team name
                    continue
                
                dept_name = row[0]
                leader_name = row[1]
                team_name = row[3]
                jira_project = row[4] or 'N/A'
                work_stream = row[5] or 'N/A'
                repo_link = row[6]
                board_link = row[7]
                focus_areas = row[8] or 'Engineering delivery'
                skills = row[9] or 'TBC'
                slack_channels = row[15]
                standup_raw = row[16]
                wiki_link = row[18]

                dept_obj = depts.get(dept_name)
                if not dept_obj:
                    # Fallback to a default department if none specified
                    dept_obj = Department.objects.get_or_create(department_name="Engineering")[0]

                # 1. Create Team
                team = Team.objects.create(
                    department=dept_obj,
                    team_name=team_name,
                    mission=focus_areas,
                    lead_email=f"{leader_name.lower().replace(' ', '.')}@sky.uk" if leader_name else "engineering@sky.uk",
                    team_leader_name=leader_name or "TBC",
                    work_stream=work_stream,
                    project_name=jira_project,
                    project_codebase=str(repo_link) if repo_link else 'N/A',
                    tech_tags=skills,
                    status='Active'
                )
                teams_created[team_name] = team

                # Helper to format links
                def fmt_link(url):
                    if not url: return None
                    url = str(url).strip()
                    if url.startswith('http'): return url
                    return f"https://{url}"

                # 2. Map Satellite Links
                if repo_link:
                    RepositoryLink.objects.create(team=team, repo_name="Primary Git Repo", repo_url=fmt_link(repo_link))
                
                if wiki_link:
                    WikiLink.objects.create(team=team, wikki_description="Team Documentation Wiki", wikki_link=fmt_link(wiki_link))
                
                if board_link:
                    BoardLink.objects.create(team=team, board_type="Jira Board", board_url=fmt_link(board_link))

                # 3. Handle Standup
                if standup_raw:
                    standup_str = str(standup_raw)
                    s_time = time(9, 30) # Default
                    s_link = ""
                    # Try to find link in string
                    parts = standup_str.split(' ')
                    for p in parts:
                        if '.' in p or 'http' in p: 
                            s_link = fmt_link(p.strip('()'))
                        if ':' in p and len(p) <= 5: 
                            try:
                                h, m = map(int, p.split(':'))
                                s_time = time(h, m)
                            except: pass
                    
                    if s_link:
                        StandupInfo.objects.create(team=team, standup_time=s_time, standup_link=s_link)

                # 4. Map Contact Channels
                if slack_channels:
                    for chan in str(slack_channels).split(','):
                        ContactChannel.objects.create(team=team, channel_type='slack', channel_value=chan.strip())
                ContactChannel.objects.create(team=team, channel_type='email', channel_value=team.lead_email)

            self.stdout.write(f'  Created {len(teams_created)} teams with satellite links.')

            # 6. Process Dependencies (Bidirectional)
            for row in rows:
                from_team_name = row[3]
                downstream_names_raw = row[10]
                
                if not from_team_name or not downstream_names_raw or downstream_names_raw == 'None':
                    continue
                
                from_team = teams_created.get(from_team_name)
                if not from_team: continue

                # Split by comma or semicolon
                downst_str = str(downstream_names_raw).replace(';', ',')
                downstream_names = [d.strip() for d in downst_str.split(',') if d.strip()]
                
                for to_name in downstream_names:
                    to_team = teams_created.get(to_name)
                    if not to_team:
                        # Partial match attempt
                        for name in teams_created:
                            if to_name.lower() in name.lower() or name.lower() in to_name.lower():
                                to_team = teams_created[name]
                                break
                    
                    if to_team and to_team != from_team:
                        Dependency.objects.get_or_create(
                            from_team=from_team, 
                            to_team=to_team, 
                            dependency_type='downstream'
                        )
                        Dependency.objects.get_or_create(
                            from_team=from_team, 
                            to_team=to_team, 
                            dependency_type='upstream'
                        )

        self.stdout.write(self.style.SUCCESS(
            f'\n Sky Engineering Registry Finalised!\n'
            f'   Teams:        {Team.objects.count()}\n'
            f'   Audit Logs:   {AuditLog.objects.count()} (System Generated)\n'
            f'   Satellite:    {RepositoryLink.objects.count()} Repos, {WikiLink.objects.count()} Wikis, {BoardLink.objects.count()} Boards\n'
            f'   Dependencies: {Dependency.objects.count()}\n'
        ))
